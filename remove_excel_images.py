"""
Excel Image Remover
This script removes all images from every sheet in an Excel file.
"""

from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import tkinter as tk
from tkinter import filedialog
import os


def select_excel_file():
    """Open a file dialog to select an Excel file."""
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    
    file_path = filedialog.askopenfilename(
        title="Select an Excel file",
        filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
    )
    
    return file_path


def remove_images_from_excel(file_path):
    """Remove all images from all sheets in the Excel file."""
    if not file_path:
        print("No file selected. Exiting...")
        return
    
    print(f"\nProcessing file: {file_path}")
    
    try:
        # Load the workbook
        workbook = load_workbook(file_path)
        
        total_images_removed = 0
        
        # Iterate through all sheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Count images in this sheet
            images_in_sheet = len(sheet._images) if hasattr(sheet, '_images') else 0
            
            if images_in_sheet > 0:
                print(f"  Sheet '{sheet_name}': Found {images_in_sheet} image(s)")
                
                # Remove all images from the sheet
                sheet._images = []
                
                total_images_removed += images_in_sheet
            else:
                print(f"  Sheet '{sheet_name}': No images found")
        
        if total_images_removed > 0:
            # Create output filename
            directory = os.path.dirname(file_path)
            filename = os.path.basename(file_path)
            name, ext = os.path.splitext(filename)
            output_path = os.path.join(directory, f"{name}_no_images{ext}")
            
            # Save the modified workbook
            workbook.save(output_path)
            
            print(f"\n✓ Successfully removed {total_images_removed} image(s) from {len(workbook.sheetnames)} sheet(s)")
            print(f"✓ Saved to: {output_path}")
        else:
            print("\n✓ No images found in any sheet. No changes made.")
        
    except Exception as e:
        print(f"\n✗ Error processing file: {str(e)}")
    
    finally:
        workbook.close()


def main():
    """Main function to run the script."""
    print("=" * 60)
    print("Excel Image Remover")
    print("=" * 60)
    print("\nThis script will remove all images from every sheet in your Excel file.")
    print("Please select an Excel file (.xlsx or .xlsm)\n")
    
    # Select file
    file_path = select_excel_file()
    
    # Process the file
    remove_images_from_excel(file_path)
    
    print("\n" + "=" * 60)
    input("\nPress Enter to exit...")


if __name__ == "__main__":
    main()
